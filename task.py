import requests
from bs4 import BeautifulSoup
import datetime
# import csv
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font

#......................HELPER FUNCTIONS.......................
def getHeaders():
    return {"accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "accept-encoding": "gzip, deflate, br",
        "accept-language": "en-GB,en;q=0.9,en-US;q=0.8,ml;q=0.7",
        "cache-control": "max-age=0",
        "dnt": "1",
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "none",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.122 Safari/537.36"}

def getUrl(stockSymbol,current,old):
    url = "https://finance.yahoo.com/quote/"+stockSymbol+".NS/history?period1="+old+"&period2="+current
    return url

def getDatetime():
    current = str(int(datetime.datetime.now().timestamp()))
    o = datetime.timedelta(days = 90)
    c = datetime.datetime.now()
    x = c-o
    old = str(int(x.timestamp()))
    return (current,old)
    
def getNameSymbol(sheet_obj,rowNo):
    cell_obj = sheet_obj['A'+str(rowNo): 'B'+str(rowNo)]
    for cell1, cell2 in cell_obj:
        return (cell1.value.replace(" ", ""), cell2.value.replace(" ", ""))

def createOutputFile(sheet):
    c1 = sheet.cell(row = 1, column = 1) 
    c1.value = "Sr No"
    c1 = sheet.cell(row = 1, column = 2) 
    c1.value = "Stock Name"
    c2 = sheet.cell(row = 1, column = 3) 
    c2.value = "Symbol"
    c3 = sheet.cell(row = 1, column = 4) 
    c3.value = "Date"
    c4 = sheet.cell(row = 1, column = 5) 
    c4.value = "Close"
    sheet.cell(row = 1, column = 1).font = Font(bold = True)
    sheet.cell(row = 1, column = 2).font = Font(bold = True)
    sheet.cell(row = 1, column = 3).font = Font(bold = True)
    sheet.cell(row = 1, column = 4).font = Font(bold = True)
    sheet.cell(row = 1, column = 5).font = Font(bold = True)
    sheet.column_dimensions['A'].width = 6
    sheet.column_dimensions['B'].width = 33.71
    sheet.column_dimensions['C'].width = 13.86
    sheet.column_dimensions['D'].width = 17
    sheet.column_dimensions['E'].width = 8.14
    
#......................HELPER FUNCTIONS.......................




#......................MAIN LOGIC.............................

#.........Read stock list..........
path = "D:/List of Stocks case study (1).xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
totalRows = sheet_obj.max_row

#.........Create output file........
wb = openpyxl.Workbook() 
sheet = wb.active 
createOutputFile(sheet)
wb.save("D:/output.xlsx")

#...........Process and append each stock data into output.xlsx file..........

for rowNo in range(2,totalRows+1,1):
    try:
        stockName,stockSymbol = getNameSymbol(sheet_obj,rowNo)
        print("("+str(rowNo-1)+")"+" Processing for " + stockName+" ("+stockSymbol+")")
        headers = getHeaders()
        current,old = getDatetime()
        url = getUrl(stockSymbol,current,old)
        response = requests.get(url,headers=headers)
        soup = BeautifulSoup(response.text, 'lxml')
        print(response.status_code)
        table1 = soup.find('table')

        #.......gettting columns.......
        headers = []
        for i in table1.find_all('th'):
            title = i.text
            headers.append(title)
        headers = [i.lower() for i in headers]
        column = {}
        if 'date' in headers and 'close*' in headers:
            column['Date'] = headers.index('date')
            column['Closing Price'] = headers.index('close*')

        #.......getting rows.......
        date = column["Date"]
        closingPrice = column["Closing Price"]
        sheet.append(['','','','',''])
        row1 = [rowNo-1,stockName,stockSymbol]
        for j in table1.find_all('tr')[1:]:
            row_data = j.find_all('td')
            if closingPrice < len(row_data):
                row = [row_data[date].text,float(row_data[closingPrice].text.replace(',',''))]
                print(row)
                sheet.append(row1+row)
                row1 = ['','','']
        wb.save("D:/output.xlsx")
    except:
        sheet.append(['','','','',''])
        sheet.append([rowNo-1,stockName,stockSymbol,'',''])
        wb.save("D:/output.xlsx")
